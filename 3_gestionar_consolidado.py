import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuración de los nombres de los consolidados y sus tablas, con rutas completas
carpeta_raiz = "C:/report_download/descargas/"
consolidados = {
    "CCM": f"{carpeta_raiz}/CCM/Consolidado_CCM.xlsx",
    "PRR": f"{carpeta_raiz}/PRR/Consolidado_PRR.xlsx",
    "CCM-ESP": f"{carpeta_raiz}/CCM-ESP/Consolidado_CCM-ESP.xlsx",
    "SOL": f"{carpeta_raiz}/SOL/Consolidado_SOL.xlsx"
}

# Columnas que deben estar en formato de fecha
fecha_columnas = ["FechaExpendiente", "FechaEtapaAprobacionMasivaFin", "FechaPre"]

def procesar_consolidados():
    for tabla_nombre, archivo in consolidados.items():
        try:
            if not os.path.exists(archivo):
                print(f"El archivo {archivo} no se encontró. Saltando...")
                continue

            print(f"Procesando: {archivo}")

            # Leer el archivo Excel
            df = pd.read_excel(archivo)

            # Optimizar: Comprobar y eliminar primera columna si está presente
            if df.columns[0].startswith("Textbox4"):
                df.drop(columns=[df.columns[0]], inplace=True)

            # Formatear columnas de fecha (vectorizado)
            for columna in fecha_columnas:
                if columna in df.columns:
                    df[columna] = pd.to_datetime(df[columna], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')

            # Añadir columnas adicionales de manera vectorizada
            if "Pre_Concluido" not in df.columns:
                df["Pre_Concluido"] = (
                    (df["EstadoTramite"] != "PENDIENTE") |
                    (df["EstadoTramite"] == "PENDIENTE") & df["EstadoPre"].notna()
                ).map({True: "SI", False: "NO"})

            if "Evaluado" not in df.columns:
                df["Evaluado"] = (
                    (df["EstadoTramite"] != "PENDIENTE") |
                    (df["EstadoTramite"] == "PENDIENTE") & 
                    (df["FechaEtapaAprobacionMasivaFin"].notna() | df["EstadoPre"].notna())
                ).map({True: "SI", False: "NO"})

            # Guardar en formato de tabla
            guardar_como_tabla(archivo, df, f"BASE_{tabla_nombre}")

        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

# Función optimizada para guardar el DataFrame como tabla
def guardar_como_tabla(archivo, df, tabla_nombre):
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        # Renombrar la pestaña si es necesario
        if ws.title != tabla_nombre:
            ws.title = tabla_nombre

        # Escribir los datos de forma optimizada
        ws.delete_cols(1, ws.max_column)  # Elimina contenido previo
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Crear tabla en OpenPyXL
        table_ref = f"A1:{chr(64 + df.shape[1])}{df.shape[0] + 1}"
        table = Table(displayName=tabla_nombre, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Guardar cambios
        wb.save(archivo)
        print(f"Archivo guardado: {archivo} con tabla {tabla_nombre}")
    except Exception as e:
        print(f"Error al guardar como tabla: {e}")

# Ejecutar
procesar_consolidados()
