import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuración de los nombres de los consolidados y sus tablas
carpeta_raiz = "C:/report_download/descargas/"
consolidados = {
    "CCM": f"{carpeta_raiz}/CCM/Consolidado_CCM.xlsx",
    "PRR": f"{carpeta_raiz}/PRR/Consolidado_PRR.xlsx",
    "CCM-ESP": f"{carpeta_raiz}/CCM-ESP/Consolidado_CCM-ESP.xlsx",
    "SOL": f"{carpeta_raiz}/SOL/Consolidado_SOL.xlsx"
}

# Columnas que deben estar en formato de fecha
fecha_columnas = ["FechaExpendiente", "FechaEtapaAprobacionMasivaFin", "FechaPre"]

# Función para validar y procesar fechas
def formatear_columnas_fecha(df, columnas):
    for columna in columnas:
        if columna in df.columns:
            df[columna] = pd.to_datetime(df[columna], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
    return df

# Función principal para procesar los consolidados
def procesar_consolidados():
    for tabla_nombre, archivo in consolidados.items():
        try:
            if not os.path.exists(archivo):
                print(f"El archivo {archivo} no se encontró. Saltando...")
                continue

            print(f"Procesando: {archivo}")

            # Leer el archivo Excel
            df = pd.read_excel(archivo)

            # Validar y eliminar columnas innecesarias
            if len(df.columns) > 0 and df.columns[0].startswith("Textbox4"):
                df.drop(columns=[df.columns[0]], inplace=True)

            # Formatear columnas de fecha
            df = formatear_columnas_fecha(df, fecha_columnas)

            # Validar columnas requeridas
            columnas_necesarias = ["EstadoTramite", "EstadoPre", "FechaEtapaAprobacionMasivaFin"]
            for columna in columnas_necesarias:
                if columna not in df.columns:
                    raise ValueError(f"La columna {columna} no está presente en el archivo {archivo}.")

            # Añadir columnas adicionales
            if "Pre_Concluido" not in df.columns:
                df["Pre_Concluido"] = (
                    (df["EstadoTramite"] != "PENDIENTE") |
                    ((df["EstadoTramite"] == "PENDIENTE") & df["EstadoPre"].notna())
                ).map({True: "SI", False: "NO"})

            if "Evaluado" not in df.columns:
                df["Evaluado"] = (
                    (df["EstadoTramite"] != "PENDIENTE") |
                    ((df["EstadoTramite"] == "PENDIENTE") & 
                    (df["FechaEtapaAprobacionMasivaFin"].notna() | df["EstadoPre"].notna()))
                ).map({True: "SI", False: "NO"})

            # Guardar en formato de tabla
            guardar_como_tabla(archivo, df, f"BASE_{tabla_nombre}")

        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

# Función para guardar el DataFrame como tabla en el archivo Excel
def guardar_como_tabla(archivo, df, tabla_nombre):
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        # Renombrar la pestaña si es necesario
        if ws.title != tabla_nombre:
            ws.title = tabla_nombre

        # Sobrescribir únicamente el rango utilizado
        max_rows = len(df) + 1
        max_cols = len(df.columns)
        for r_idx in range(1, max_rows + 1):
            for c_idx in range(1, max_cols + 1):
                ws.cell(row=r_idx, column=c_idx).value = None

        # Escribir el DataFrame en el archivo Excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Crear tabla en OpenPyXL
        table_ref = f"A1:{chr(64 + df.shape[1])}{df.shape[0] + 1}"
        table = Table(displayName=tabla_nombre, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Guardar cambios
        wb.save(archivo)
        print(f"Archivo guardado: {archivo} con tabla {tabla_nombre}")
    except Exception as e:
        print(f"Error al guardar como tabla: {e}")
