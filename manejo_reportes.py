import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def procesar_carpeta(carpeta, nombre_hoja):
    """
    Procesa una carpeta específica usando pandas y openpyxl en lugar de win32com.
    
    Args:
        carpeta: Nombre de la carpeta a procesar
        nombre_hoja: Nombre de la hoja de consolidado
    """
    # Configuración y rutas manteniendo la estructura original
    current_dir = os.path.dirname(os.path.abspath(__file__))
    carpeta_path = os.path.join(current_dir, "manejo_reportes", carpeta)
    archivo_pend = os.path.join(carpeta_path, f"{carpeta}PEND.xlsx")
    archivo_asignaciones = os.path.join(carpeta_path, "ASIGNACIONES.xlsx")

    if not os.path.exists(archivo_pend) or not os.path.exists(archivo_asignaciones):
        print(f"Archivos faltantes en {carpeta}.")
        return

    try:
        print(f"Procesando carpeta: {carpeta}...")

        # Leer archivo de pendientes
        df_pend = pd.read_excel(archivo_pend, header=3)  # Empezar desde la fila 4
        
        # Identificar las columnas según la carpeta
        if carpeta == "SOL":
            col_tramite_idx, col_nombre_idx = 5, 23  # F = 5, X = 23
        else:
            col_tramite_idx, col_nombre_idx = 5, 32  # F = 5, AG = 32

        # Obtener las columnas por índice
        columnas = df_pend.columns.tolist()
        col_tramite = columnas[col_tramite_idx]
        col_nombre = columnas[col_nombre_idx]
        
        # Filtrar datos de pendientes
        df_pend = df_pend[[col_tramite, col_nombre]].copy()
        df_pend.columns = ["EXPEDIENTE", "EVALUADOR"]
        
        # Filtrar trámites válidos y limpiar datos
        df_pend = df_pend.dropna()
        mask = df_pend["EXPEDIENTE"].astype(str).str.match(r'^(LM|LS|MR|LN).*')
        datos_filtrados = df_pend[mask].values.tolist()

        print(f"Registros filtrados en {carpeta}: {len(datos_filtrados)}")

        # Trabajar con el archivo de asignaciones
        wb_asignaciones = load_workbook(archivo_asignaciones)
        
        # Limpiar REPORTADO_SIM
        if "REPORTADO_SIM" in wb_asignaciones.sheetnames:
            ws_reportado = wb_asignaciones["REPORTADO_SIM"]
            for row in ws_reportado.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

        # Escribir datos filtrados en REPORTADO_SIM
        ws_reportado = wb_asignaciones["REPORTADO_SIM"]
        for i, (exp, eval) in enumerate(datos_filtrados, start=2):
            ws_reportado.cell(row=i, column=1, value=exp)
            ws_reportado.cell(row=i, column=2, value=eval)

        # Leer datos de IDENTIFICADOS
        ws_identificados = wb_asignaciones["IDENTIFICADOS"]
        datos_identificados = []
        for row in ws_identificados.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Solo si hay datos en ambas columnas
                datos_identificados.append([row[0], row[1]])

        # Crear o reemplazar hoja CONSOLIDADO
        if nombre_hoja in wb_asignaciones.sheetnames:
            del wb_asignaciones[nombre_hoja]
        ws_consolidado = wb_asignaciones.create_sheet(nombre_hoja)

        # Escribir encabezados
        ws_consolidado.cell(row=1, column=1, value="EXPEDIENTE")
        ws_consolidado.cell(row=1, column=2, value="EVALUADOR")

        # Combinar y escribir datos
        datos_consolidados = datos_identificados + datos_filtrados
        for i, (exp, eval) in enumerate(datos_consolidados, start=2):
            ws_consolidado.cell(row=i, column=1, value=exp)
            ws_consolidado.cell(row=i, column=2, value=eval)

        # Crear tabla
        tab = Table(displayName=nombre_hoja.replace("-", "_"),
                   ref=f"A1:B{len(datos_consolidados) + 1}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws_consolidado.add_table(tab)

        # Guardar cambios
        wb_asignaciones.save(archivo_asignaciones)
        print(f"Consolidado creado en {carpeta}.")

    except Exception as e:
        print(f"Error procesando la carpeta {carpeta}: {str(e)}")
        import traceback
        print(traceback.format_exc())

def manejar_reportes():
    """
    Función principal para manejar los reportes.
    """
    print("\n>>> Ejecutando: Manejo de reportes evaluados")
    
    try:
        # Procesar cada carpeta
        for carpeta, nombre_hoja in [
            ("CCM", "CONSOLIDADO_CCM_X_EVAL"),
            ("PRR", "CONSOLIDADO_PRR_X_EVAL"),
            ("SOL", "CONSOLIDADO_SOL_X_EVAL")
        ]:
            procesar_carpeta(carpeta, nombre_hoja)

    except Exception as e:
        print(f"Error en el proceso: {str(e)}")
        import traceback
        print(traceback.format_exc())

    print("Proceso completado.")

if __name__ == "__main__":
    manejar_reportes()