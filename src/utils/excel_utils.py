import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_excel_table(writer, df, sheet_name, title=None):
    """
    Aplica formato profesional a una tabla de Excel.
    
    Args:
        writer: ExcelWriter object
        df: DataFrame a escribir
        sheet_name: Nombre de la hoja
        title: Título opcional para la tabla
    """
    # Escribir DataFrame
    df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1 if title else 0)
    
    # Obtener la hoja de trabajo
    worksheet = writer.sheets[sheet_name]
    
    # Estilos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Si hay título, agregarlo
    if title:
        worksheet.merge_cells(f'A1:{get_column_letter(df.shape[1] + 1)}1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
    
    # Formato de encabezados
    start_row = 2 if title else 1
    for col in range(1, df.shape[1] + 2):  # +2 para incluir la columna de índice
        cell = worksheet.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Formato de celdas de datos
    for row in range(start_row + 1, df.shape[0] + start_row + 1):
        for col in range(1, df.shape[1] + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    for col in range(1, df.shape[1] + 2):
        worksheet.column_dimensions[get_column_letter(col)].width = 15

def create_excel_download(df, filename, sheet_name="Reporte", title=None):
    """
    Crea un archivo Excel formateado para descarga.
    
    Args:
        df: DataFrame a exportar
        filename: Nombre del archivo
        sheet_name: Nombre de la hoja
        title: Título opcional para la tabla
    
    Returns:
        BytesIO object con el archivo Excel
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        format_excel_table(writer, df, sheet_name, title)
    output.seek(0)
    return output 